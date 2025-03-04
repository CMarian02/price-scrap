import pandas as pd
from openpyxl import load_workbook

def get_products_from_excel(file_path):
    df = pd.read_excel(file_path)
    df = df.dropna(subset=[df.columns[0], df.columns[1]])
    products = pd.Series(df.iloc[:, 1].values, index=df.iloc[:, 0]).to_dict()
    return products

def add_product_data(file_path:str, product_name: str, product_price: str, date_time: str): 
    
    wb = load_workbook(file_path)
    ws = wb.active
    last_row = ws.max_row + 1
    ws[f'A{last_row}'] = product_name
    ws[f'B{last_row}'] = product_price
    ws[f'C{last_row}'] = date_time

    wb.save(file_path)
    